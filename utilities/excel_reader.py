import os
from openpyxl import load_workbook


def read_excel_data(file_name, sheet_name):

    project_root = os.path.dirname(os.path.dirname(__file__))

    file_path = os.path.join(project_root, "test_data", file_name)

    workbook = load_workbook(file_path)

    sheet = workbook[sheet_name]

    data = []

    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):

        row_data = dict(zip(headers, row))

        data.append(row_data)

    workbook.close()

    return data